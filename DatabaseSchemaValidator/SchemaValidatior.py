import os
import re
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import pyodbc
import psycopg2
from config import SQL_SERVER_CONFIG, POSTGRES_CONFIG, DB_LIST
from mappings import PROCEDURE_NAME_MAP, EVENT_TRIGGER_NAME_MAP

def get_sqlserver_connection():
    # Add support for Windows Authentication if 'windows_auth' key is True in config
    if SQL_SERVER_CONFIG.get('windows_auth', False):
        conn_str = (
            f"DRIVER={SQL_SERVER_CONFIG['driver']};"
            f"SERVER={SQL_SERVER_CONFIG['server']};"
            f"DATABASE={SQL_SERVER_CONFIG['database']};"
            f"Trusted_Connection=yes;"
        )
    else:
        conn_str = (
            f"DRIVER={SQL_SERVER_CONFIG['driver']};"
            f"SERVER={SQL_SERVER_CONFIG['server']};"
            f"DATABASE={SQL_SERVER_CONFIG['database']};"
            f"UID={SQL_SERVER_CONFIG['username']};"
            f"PWD={SQL_SERVER_CONFIG['password']}"
        )
    return pyodbc.connect(conn_str)

def get_postgres_connection():
    return psycopg2.connect(**POSTGRES_CONFIG)

# --- Extraction stubs (to be filled in) ---
def extract_tables(conn, dbtype):
    cursor = conn.cursor()
    if dbtype == 'sql':
        cursor.execute("""
            SELECT TABLE_SCHEMA, TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE = 'BASE TABLE'
        """)
        return [{'schema': row[0], 'name': row[1], 'fullname': f"{row[0]}.{row[1]}", 'dbtype': 'sql'} for row in cursor.fetchall()]
    else:
        cursor.execute("""
            SELECT table_schema, table_name FROM information_schema.tables WHERE table_type = 'BASE TABLE' AND table_schema NOT IN ('pg_catalog', 'information_schema')
        """)
        return [{'schema': row[0], 'name': row[1], 'fullname': f"{row[0]}.{row[1]}", 'dbtype': 'pg'} for row in cursor.fetchall()]

def extract_columns(conn, dbtype):
    cursor = conn.cursor()
    if dbtype == 'sql':
        cursor.execute("""
            SELECT TABLE_SCHEMA, TABLE_NAME, COLUMN_NAME, DATA_TYPE, IS_NULLABLE, COLUMN_DEFAULT
            FROM INFORMATION_SCHEMA.COLUMNS
        """)
        return [
            {'schema': row[0], 'table': row[1], 'name': row[2], 'datatype': row[3], 'nullable': row[4], 'default': row[5], 'fullname': f"{row[0]}.{row[1]}", 'dbtype': 'sql'}
            for row in cursor.fetchall()
        ]
    else:
        cursor.execute("""
            SELECT table_schema, table_name, column_name, data_type, is_nullable, column_default
            FROM information_schema.columns WHERE table_schema NOT IN ('pg_catalog', 'information_schema')
        """)
        return [
            {'schema': row[0], 'table': row[1], 'name': row[2], 'datatype': row[3], 'nullable': row[4], 'default': row[5], 'fullname': f"{row[0]}.{row[1]}", 'dbtype': 'pg'}
            for row in cursor.fetchall()
        ]

def extract_constraints(conn, dbtype):
    cursor = conn.cursor()
    constraints = []
    if dbtype == 'sql':
        # PK, FK, Unique, Check, Default
        cursor.execute("""
            SELECT tc.TABLE_SCHEMA, tc.TABLE_NAME, tc.CONSTRAINT_NAME, tc.CONSTRAINT_TYPE
            FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS tc
        """)
        for row in cursor.fetchall():
            constraints.append({'schema': row[0], 'table': row[1], 'name': row[2], 'type': row[3], 'fullname': f"{row[0]}.{row[1]}", 'dbtype': 'sql'})
        # Foreign keys: add referenced table/columns to definition
        cursor.execute("""
            SELECT fk.CONSTRAINT_SCHEMA, fk.TABLE_NAME, fk.CONSTRAINT_NAME, 'FOREIGN KEY', 
                STUFF((SELECT ',' + kcu.COLUMN_NAME
                       FROM INFORMATION_SCHEMA.KEY_COLUMN_USAGE kcu
                       WHERE kcu.CONSTRAINT_NAME = fk.CONSTRAINT_NAME AND kcu.TABLE_NAME = fk.TABLE_NAME
                       ORDER BY kcu.ORDINAL_POSITION
                       FOR XML PATH('')), 1, 1, '') AS FK_COLUMNS,
                rc.UNIQUE_CONSTRAINT_SCHEMA, rc.UNIQUE_CONSTRAINT_NAME, 
                (SELECT TOP 1 kcu2.TABLE_NAME FROM INFORMATION_SCHEMA.KEY_COLUMN_USAGE kcu2 WHERE kcu2.CONSTRAINT_NAME = rc.UNIQUE_CONSTRAINT_NAME) AS REF_TABLE,
                STUFF((SELECT ',' + kcu2.COLUMN_NAME
                       FROM INFORMATION_SCHEMA.KEY_COLUMN_USAGE kcu2
                       WHERE kcu2.CONSTRAINT_NAME = rc.UNIQUE_CONSTRAINT_NAME
                       ORDER BY kcu2.ORDINAL_POSITION
                       FOR XML PATH('')), 1, 1, '') AS REF_COLUMNS
            FROM INFORMATION_SCHEMA.REFERENTIAL_CONSTRAINTS rc
            JOIN INFORMATION_SCHEMA.TABLE_CONSTRAINTS fk ON rc.CONSTRAINT_NAME = fk.CONSTRAINT_NAME
        """)
        for row in cursor.fetchall():
            fk_cols = row[4]
            ref_table = row[7]
            ref_cols = row[8]
            definition = f"FOREIGN KEY ({{fk_cols}}) REFERENCES {{ref_table}} ({{ref_cols}})"
            constraints.append({'schema': row[0], 'table': row[1], 'name': row[2], 'type': row[3], 'definition': definition, 'fullname': f"{row[0]}.{row[1]}", 'dbtype': 'sql'})
        # Check constraints (fix: get table name from CONSTRAINT_TABLE_USAGE)
        cursor.execute("""
            SELECT cc.CONSTRAINT_SCHEMA, ctu.TABLE_NAME, cc.CONSTRAINT_NAME, 'CHECK', cc.CHECK_CLAUSE
            FROM INFORMATION_SCHEMA.CHECK_CONSTRAINTS cc
            JOIN INFORMATION_SCHEMA.CONSTRAINT_TABLE_USAGE ctu ON cc.CONSTRAINT_NAME = ctu.CONSTRAINT_NAME
        """)
        for row in cursor.fetchall():
            constraints.append({'schema': row[0], 'table': row[1], 'name': row[2], 'type': row[3], 'definition': row[4], 'fullname': f"{row[0]}.{row[1]}", 'dbtype': 'sql'})
        # Default constraints
        cursor.execute("""
            SELECT c.TABLE_SCHEMA, c.TABLE_NAME, c.COLUMN_NAME, 'DEFAULT', c.COLUMN_DEFAULT
            FROM INFORMATION_SCHEMA.COLUMNS c WHERE c.COLUMN_DEFAULT IS NOT NULL
        """)
        for row in cursor.fetchall():
            constraints.append({'schema': row[0], 'table': row[1], 'name': row[2], 'type': row[3], 'definition': row[4], 'fullname': f"{row[0]}.{row[1]}", 'dbtype': 'sql'})
        # Synthesize NOT NULL constraints for each column
        cursor.execute("""
            SELECT TABLE_SCHEMA, TABLE_NAME, COLUMN_NAME, IS_NULLABLE
            FROM INFORMATION_SCHEMA.COLUMNS
        """)
        for row in cursor.fetchall():
            if row[3].strip().upper() == 'NO':
                # Synthesize a check constraint for NOT NULL
                constraints.append({
                    'schema': row[0],
                    'table': row[1],
                    'name': f'not_null_{row[2]}',
                    'type': 'CHECK',
                    'definition': f'([{row[2]}] IS NOT NULL)',
                    'fullname': f"{row[0]}.{row[1]}",
                    'dbtype': 'sql',
                    'synthesized': True
                })
    else:
        cursor.execute("""
            SELECT tc.table_schema, tc.table_name, tc.constraint_name, tc.constraint_type
            FROM information_schema.table_constraints tc WHERE tc.table_schema NOT IN ('pg_catalog', 'information_schema')
        """)
        for row in cursor.fetchall():
            constraints.append({'schema': row[0], 'table': row[1], 'name': row[2], 'type': row[3], 'fullname': f"{row[0]}.{row[1]}", 'dbtype': 'pg'})
        # Check constraints
        cursor.execute("""
            SELECT cc.constraint_schema, ctu.table_name, cc.constraint_name, 'CHECK', cc.check_clause
            FROM information_schema.check_constraints cc
            JOIN information_schema.constraint_table_usage ctu ON cc.constraint_name = ctu.constraint_name
            WHERE cc.constraint_schema NOT IN ('pg_catalog', 'information_schema')
        """)
        for row in cursor.fetchall():
            constraints.append({'schema': row[0], 'table': row[1], 'name': row[2], 'type': row[3], 'definition': row[4], 'fullname': f"{row[0]}.{row[1]}", 'dbtype': 'pg'})
        # Default constraints
        cursor.execute("""
            SELECT table_schema, table_name, column_name, 'DEFAULT', column_default
            FROM information_schema.columns WHERE column_default IS NOT NULL AND table_schema NOT IN ('pg_catalog', 'information_schema')
        """)
        for row in cursor.fetchall():
            # Synthesize a definition for PG default constraints
            definition = f"DEFAULT ({row[4]}) FOR {row[2]}"
            constraints.append({'schema': row[0], 'table': row[1], 'name': row[2], 'type': row[3], 'definition': definition, 'fullname': f"{row[0]}.{row[1]}", 'dbtype': 'pg'})
    return constraints

def extract_indexes(conn, dbtype):
    cursor = conn.cursor()
    indexes = []
    if dbtype == 'sql':
        cursor.execute("""
            SELECT
                t.name AS TableName,
                i.name AS IndexName,
                c.name AS ColumnName,
                i.type_desc AS IndexType,
                s.name AS SchemaName
            FROM 
                sys.indexes i
            JOIN 
                sys.index_columns ic ON i.object_id = ic.object_id AND i.index_id = ic.index_id
            JOIN 
                sys.columns c ON ic.object_id = c.object_id AND ic.column_id = c.column_id
            JOIN 
                sys.tables t ON i.object_id = t.object_id
            JOIN 
                sys.schemas s ON t.schema_id = s.schema_id
            WHERE 
                i.is_primary_key = 0 AND i.is_unique_constraint = 0 AND i.type_desc <> 'HEAP'
            ORDER BY 
                s.name, t.name, i.name, ic.key_ordinal
        """)
        index_map = {}
        for row in cursor.fetchall():
            table = row[0]
            index = row[1]
            column = row[2]
            idx_type = row[3]
            schema = row[4]
            key = (schema, table, index, idx_type)
            if key not in index_map:
                index_map[key] = []
            index_map[key].append(column)
        for (schema, table, index, idx_type), columns in index_map.items():
            indexes.append({
                'schema': schema,
                'table': table,
                'name': index,
                'type': idx_type,
                'columns': ','.join(columns),
                'fullname': f"{schema}.{table}",
                'dbtype': 'sql'
            })
    else:
        cursor.execute("""
            SELECT schemaname, tablename, indexname, indexdef
            FROM pg_indexes 
            WHERE schemaname NOT IN ('pg_catalog', 'information_schema')
        """)
        for row in cursor.fetchall():
            # Exclude primary key indexes by name (any name starting with 'pk') and definition
            index_name_lower = (row[2] or '').lower()
            if 'primary key' in row[3].lower() or index_name_lower.startswith('pk'):
                continue
            m = re.search(r'\(([^)]+)\)', row[3])
            columns = m.group(1).replace(' ', '') if m else ''
            idx_type = 'UNIQUE' if 'unique' in row[3].lower() else 'INDEX'
            indexes.append({'schema': row[0], 'table': row[1], 'name': row[2], 'type': idx_type, 'definition': row[3], 'columns': columns, 'fullname': f"{row[0]}.{row[1]}", 'dbtype': 'pg'})
    return indexes

def extract_triggers(conn, dbtype):
    cursor = conn.cursor()
    triggers = []
    if dbtype == 'sql':
        cursor.execute("""
            SELECT s.name, t.name, tr.name
            FROM sys.triggers tr
            JOIN sys.tables t ON tr.parent_id = t.object_id
            JOIN sys.schemas s ON t.schema_id = s.schema_id
        """)
        for row in cursor.fetchall():
            triggers.append({'schema': row[0], 'table': row[1], 'name': row[2], 'fullname': f"{row[0]}.{row[1]}", 'dbtype': 'sql'})
    else:
        cursor.execute("""
            SELECT event_object_schema, event_object_table, trigger_name
            FROM information_schema.triggers WHERE event_object_schema NOT IN ('pg_catalog', 'information_schema')
        """)
        for row in cursor.fetchall():
            triggers.append({'schema': row[0], 'table': row[1], 'name': row[2], 'fullname': f"{row[0]}.{row[1]}", 'dbtype': 'pg'})
    return triggers

def extract_event_triggers(conn, dbtype):
    triggers = []
    if dbtype == 'pg':
        cursor = conn.cursor()
        cursor.execute("""
            SELECT evtname FROM pg_event_trigger
        """)
        triggers = [{'name': row[0], 'dbtype': 'pg'} for row in cursor.fetchall()]
    elif dbtype == 'sql':
        # SQL Server does not have event triggers like PG, but for completeness, try to get DDL triggers
        cursor = conn.cursor()
        cursor.execute("""
            SELECT name FROM sys.triggers WHERE parent_class = 0
        """)
        triggers = [{'name': row[0], 'dbtype': 'sql'} for row in cursor.fetchall()]
    return triggers

def extract_views(conn, dbtype):
    cursor = conn.cursor()
    if dbtype == 'sql':
        cursor.execute("""
            SELECT TABLE_SCHEMA, TABLE_NAME FROM INFORMATION_SCHEMA.VIEWS
        """)
        return [{'schema': row[0], 'name': row[1], 'fullname': f"{row[0]}.{row[1]}", 'dbtype': 'sql'} for row in cursor.fetchall()]
    else:
        cursor.execute("""
            SELECT table_schema, table_name FROM information_schema.views WHERE table_schema NOT IN ('pg_catalog', 'information_schema')
        """)
        return [{'schema': row[0], 'name': row[1], 'fullname': f"{row[0]}.{row[1]}", 'dbtype': 'pg'} for row in cursor.fetchall()]

def extract_functions(conn, dbtype):
    cursor = conn.cursor()
    functions = []
    if dbtype == 'sql':
        cursor.execute("""
            SELECT ROUTINE_SCHEMA, ROUTINE_NAME, ROUTINE_TYPE
            FROM INFORMATION_SCHEMA.ROUTINES
            WHERE ROUTINE_TYPE = 'FUNCTION'
        """)
        for row in cursor.fetchall():
            # SQL Server doesn't distinguish trigger functions, so mark as 'normal'
            functions.append({'schema': row[0], 'name': row[1], 'type': row[2].lower(), 'function_type': 'normal', 'fullname': f"{row[0]}.{row[1]}", 'dbtype': 'sql'})
    else:
        cursor.execute("""
            SELECT routine_schema, routine_name, routine_type, data_type
            FROM information_schema.routines
            WHERE routine_schema NOT IN ('pg_catalog', 'information_schema')
              AND routine_type = 'FUNCTION'
        """)
        for row in cursor.fetchall():
            # Use data_type to classify trigger functions
            func_type = 'trigger' if row[3] and row[3].lower() in ('trigger', 'event_trigger') else 'normal'
            functions.append({'schema': row[0], 'name': row[1], 'type': row[2].lower(), 'function_type': func_type, 'fullname': f"{row[0]}.{row[1]}", 'dbtype': 'pg'})
    return functions

def extract_types(conn, dbtype):
    cursor = conn.cursor()
    types = []
    if dbtype == 'sql':
        cursor.execute("""
            SELECT s.name, t.name, t.is_table_type
            FROM sys.types t JOIN sys.schemas s ON t.schema_id = s.schema_id
            WHERE t.is_user_defined = 1
        """)
        for row in cursor.fetchall():
            types.append({
                'schema': row[0],
                'type_name': row[1],
                'type_kind': 'table' if row[2] else 'user-defined'
            })
    else:
        cursor.execute("""
            SELECT t.typname AS type_name,
                   CASE t.typtype
                        WHEN 'c' THEN 'composite'
                        WHEN 'd' THEN 'domain'
                        WHEN 'e' THEN 'enum'
                        WHEN 'r' THEN 'range'
                   END AS type_kind,
                   n.nspname AS schema
            FROM pg_type t
            JOIN pg_namespace n ON n.oid = t.typnamespace
            WHERE n.nspname IN ('dbo', 'meta', 'public')
              AND t.typtype IN ('c', 'd', 'e', 'r')
              AND (t.typrelid = 0 OR NOT EXISTS (
                    SELECT 1
                    FROM pg_class c
                    WHERE c.oid = t.typrelid
                      AND c.relkind IN ('r', 'v', 'm')
                  ))
            ORDER BY t.typname;
        """)
        for row in cursor.fetchall():
            types.append({
                'schema': row[2],
                'type_name': row[0],
                'type_kind': row[1]
            })
    return types

def extract_procedures(conn, dbtype):
    cursor = conn.cursor()
    if dbtype == 'sql':
        cursor.execute("""
            SELECT SPECIFIC_SCHEMA, SPECIFIC_NAME FROM INFORMATION_SCHEMA.ROUTINES WHERE ROUTINE_TYPE = 'PROCEDURE'
        """)
        return [{'schema': row[0], 'name': row[1], 'fullname': f"{row[0]}.{row[1]}", 'dbtype': 'sql'} for row in cursor.fetchall()]
    else:
        cursor.execute("""
            SELECT routine_schema, routine_name FROM information_schema.routines WHERE routine_type = 'PROCEDURE' AND routine_schema NOT IN ('pg_catalog', 'information_schema')
        """)
        return [{'schema': row[0], 'name': row[1], 'fullname': f"{row[0]}.{row[1]}", 'dbtype': 'pg'} for row in cursor.fetchall()]

def extract_table_counts(conn, dbtype):
    cursor = conn.cursor()
    if dbtype == 'sql':
        cursor.execute("""
            SELECT TABLE_SCHEMA, TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE = 'BASE TABLE'
        """)
        tables = cursor.fetchall()
        counts = []
        for row in tables:
            try:
                cursor.execute(f"SELECT COUNT(*) FROM [{row[0]}].[{row[1]}]")
                cnt = cursor.fetchone()[0]
            except:
                cnt = ''
            counts.append({'schema': row[0], 'name': row[1], 'fullname': f"{row[0]}.{row[1]}", 'count': cnt, 'dbtype': 'sql'})
        return counts
    else:
        cursor.execute("""
            SELECT table_schema, table_name FROM information_schema.tables WHERE table_type = 'BASE TABLE' AND table_schema NOT IN ('pg_catalog', 'information_schema')
        """)
        tables = cursor.fetchall()
        counts = []
        for row in tables:
            try:
                cursor.execute(f'SELECT COUNT(*) FROM "{row[0]}"."{row[1]}"')
                cnt = cursor.fetchone()[0]
            except:
                cnt = ''
            counts.append({'schema': row[0], 'name': row[1], 'fullname': f"{row[0]}.{row[1]}", 'count': cnt, 'dbtype': 'pg'})
        return counts

def normalize_name(name):
    return (name or '').strip().lower()

def normalize_fullname(schema, name):
    return f"{normalize_name(schema)}.{normalize_name(name)}"

def normalize_index_name(name):
    # Remove common prefixes, underscores, and lowercase
    name = (name or '').lower()
    name = re.sub(r'^(ix_|idxn|idx|pk_|pk__|uq_|uq__|ak_|ak__|unique_)', '', name)
    name = name.replace('_', '')
    return name

def normalize_index_columns(columns):
    # Normalize and sort columns for comparison
    if not columns:
        return []
    return sorted([c.strip().lower() for c in columns.split(',') if c.strip()])

def normalize_check_name(name):
    # For Postgres, strip trailing _<digits> for check constraints
    if name is None:
        return ''
    return re.sub(r'_[0-9]+$', '', name.lower())

def normalize_constraint_name(name):
    # Lowercase, remove prefixes (chk_, ck_, etc.), and strip trailing _digits
    if not name:
        return ''
    name = name.lower()
    name = re.sub(r'^(chk_|ck_|df_|default_)+', '', name)
    name = re.sub(r'_[0-9]+$', '', name)
    return name

# Exclude schemas before processing
EXCLUDED_SCHEMAS = {'aws_sqlserver_ext', 'aws_sqlserver_ext_data'}

def filter_excluded(entities):
    return [e for e in entities if normalize_name(e.get('schema','')) not in EXCLUDED_SCHEMAS]

SQL_TO_PG_TYPE_MAP = {
    'int': ['integer', 'int4'],
    'bigint': ['bigint', 'int8'],
    'smallint': ['smallint', 'int2'],
    'tinyint': ['smallint', 'int2', 'integer'],  # integer added for tinyint
    'bit': ['boolean', 'bool'],
    'varchar': ['character varying', 'varchar', 'text', 'name'],
    'nvarchar': ['character varying', 'varchar', 'text', 'name'],
    'char': ['character', 'char', 'name'],
    'nchar': ['character', 'char', 'name'],
    'text': ['text', 'character varying', 'name'],
    'ntext': ['text', 'character varying', 'name'],
    'datetime': ['timestamp', 'timestamp without time zone', 'timestamp with time zone'],
    'datetime2': ['timestamp', 'timestamp without time zone', 'timestamp with time zone'],
    'smalldatetime': ['timestamp', 'timestamp without time zone'],
    'date': ['date'],
    'time': ['time', 'time without time zone', 'time with time zone'],
    'float': ['double precision', 'float8'],
    'real': ['real', 'float4'],
    'decimal': ['numeric', 'decimal'],
    'numeric': ['numeric', 'decimal'],
    'money': ['numeric', 'decimal'],
    'smallmoney': ['numeric', 'decimal'],
    'uniqueidentifier': ['uuid'],
    'xml': ['xml', 'text'],
    'varbinary': ['bytea'],
    'binary': ['bytea'],
    'image': ['bytea'],
    'json': ['json', 'jsonb'],
    'jsonb': ['jsonb', 'json'],
}

def are_types_compatible(sql_type, pg_type):
    sql_type = (sql_type or '').lower()
    pg_type = (pg_type or '').lower()
    if sql_type == '' or pg_type == '':
        return True
    if sql_type == pg_type:
        return True
    if sql_type in SQL_TO_PG_TYPE_MAP:
        return pg_type in SQL_TO_PG_TYPE_MAP[sql_type]
    return False

def are_index_names_equivalent(sql_idx, pg_idx, table):
    sql_idx = sql_idx or ''
    pg_idx = pg_idx or ''
    table = table or ''
    norm_pg = pg_idx.lower()
    norm_sql = sql_idx.lower()
    if norm_sql == norm_pg:
        return True
    if norm_sql in norm_pg or norm_pg in norm_sql:
        return True
    expected_pg = f"ix_{table.lower()}_{norm_sql}"
    if expected_pg == norm_pg:
        return True
    if norm_pg.startswith(expected_pg[:63]):
        return True
    if norm_pg.endswith(norm_sql):
        return True
    return False

def get_entity_config(entity):
    # Returns (key_fields, compare_fields, output_columns)
    if entity == 'Tables':
        return (
            ['name'],
            [],
            ['name', 'SQL_dbtype', 'PG_dbtype', 'Status']
        )
    elif entity == 'Columns':
        return (
            ['table', 'name'],
            ['datatype', 'nullable', 'default'],
            ['table', 'name', 'SQL_datatype', 'PG_datatype', 'SQL_nullable', 'PG_nullable', 'SQL_default', 'PG_default', 'Status']
        )
    elif entity == 'Constraints':
        return (
            ['table', 'name', 'type'],
            ['type', 'definition'],
            ['table', 'name', 'type', 'SQL_definition', 'PG_definition', 'Status']
        )
    elif entity == 'Indexes':
        return (
            ['table', 'name'],
            ['name', 'type'],
            ['table', 'name', 'SQL_type', 'PG_type', 'Status']
        )
    elif entity == 'Triggers':
        return (
            ['table', 'name'],
            [],
            ['table', 'name', 'Status']
        )
    elif entity == 'EventTriggers':
        return (
            ['name'],
            [],
            ['name', 'dbtype']
        )
    elif entity == 'Views':
        return (
            ['name'],
            [],
            ['name', 'Status']
        )
    elif entity == 'Functions':
        return (
            ['name'],
            [],
            ['name', 'Status']
        )
    elif entity == 'Types':
        return (
            ['name'],
            [],
            ['name', 'Status']
        )
    elif entity == 'Procedures':
        return (
            ['name'],
            [],
            ['name', 'Status']
        )
    elif entity == 'DataCounts':
        return (
            ['name'],
            ['count'],
            ['name', 'SQL_count', 'PG_count', 'Status']
        )
    else:
        return (['name'], [], ['name', 'Status'])

def robust_index_match(sql_indexes, pg_indexes):
    # Returns a mapping of sql_key -> pg_key for best matches
    matches = {}
    used_pg = set()
    for sql_idx in sql_indexes:
        sql_table = normalize_name(sql_idx.get('table',''))
        sql_name = normalize_name(sql_idx.get('name',''))
        best_pg = None
        for pg_idx in pg_indexes:
            if id(pg_idx) in used_pg:
                continue
            pg_table = normalize_name(pg_idx.get('table',''))
            pg_name = normalize_name(pg_idx.get('name',''))
            # Table must match
            if sql_table != pg_table:
                continue
            # Robust name match: contains, prefix, suffix, etc.
            if sql_name == pg_name or sql_name in pg_name or pg_name in sql_name:
                best_pg = pg_idx
                break
            if pg_name.startswith(sql_name) or pg_name.endswith(sql_name):
                best_pg = pg_idx
                break
            if sql_name.startswith(pg_name) or sql_name.endswith(pg_name):
                best_pg = pg_idx
                break
        if best_pg:
            matches[id(sql_idx)] = id(best_pg)
            used_pg.add(id(best_pg))
    return matches

def robust_trigger_match(sql_triggers, pg_triggers):
    matches = {}
    used_pg = set()
    for sql_tr in sql_triggers:
        sql_table = normalize_name(sql_tr.get('table',''))
        sql_name = normalize_name(sql_tr.get('name',''))
        best_pg = None
        for pg_tr in pg_triggers:
            if id(pg_tr) in used_pg:
                continue
            pg_table = normalize_name(pg_tr.get('table',''))
            pg_name = normalize_name(pg_tr.get('name',''))
            if sql_table != pg_table:
                continue
            # Contains/robust match
            if sql_name == pg_name or sql_name in pg_name or pg_name in sql_name:
                best_pg = pg_tr
                break
            if pg_name.startswith(sql_name) or pg_name.endswith(sql_name):
                best_pg = pg_tr
                break
            if sql_name.startswith(pg_name) or sql_name.endswith(pg_name):
                best_pg = pg_tr
                break
        if best_pg:
            matches[id(sql_tr)] = id(best_pg)
            used_pg.add(id(best_pg))
    return matches

def match_by_keys(sql_list, pg_list, keys, extra_matchers=None):
    """Match PG records to SQL records by keys, with optional extra matchers for fuzzy logic."""
    matches = {}
    used_pg = set()
    for sql in sql_list:
        sql_key = tuple(normalize_name(sql.get(k, '')) for k in keys)
        best_pg = None
        for i, pg in enumerate(pg_list):
            if i in used_pg:
                continue
            pg_key = tuple(normalize_name(pg.get(k, '')) for k in keys)
            if sql_key == pg_key:
                best_pg = i
                break
            # Extra matchers for fuzzy/robust logic
            if extra_matchers:
                for matcher in extra_matchers:
                    if matcher(sql, pg):
                        best_pg = i
                        break
            if best_pg is not None:
                break
        if best_pg is not None:
            matches[id(sql)] = best_pg
            used_pg.add(best_pg)
    return matches

def parse_fk_details(definition):
    # Try to extract local columns and referenced table/columns from FK definition string
    # Works for both SQL and PG definition formats
    if not definition:
        return [], '', []
    # Example: FOREIGN KEY (col1, col2) REFERENCES ref_table (ref_col1, ref_col2)
    m = re.search(r'FOREIGN KEY \(([^)]+)\) REFERENCES ([^ (]+) \(([^)]+)\)', definition, re.IGNORECASE)
    if m:
        local_cols = [c.strip().lower() for c in m.group(1).split(',')]
        ref_table = m.group(2).strip().lower()
        ref_cols = [c.strip().lower() for c in m.group(3).split(',')]
        return local_cols, ref_table, ref_cols
    # Try PG style: REFERENCES ref_table(col)
    m = re.search(r'REFERENCES ([^ (]+)\(([^)]+)\)', definition, re.IGNORECASE)
    if m:
        ref_table = m.group(1).strip().lower()
        ref_cols = [c.strip().lower() for c in m.group(2).split(',')]
        # Try to get local columns (may not be present)
        m2 = re.search(r'FOREIGN KEY \(([^)]+)\)', definition, re.IGNORECASE)
        if m2:
            local_cols = [c.strip().lower() for c in m2.group(1).split(',')]
        else:
            local_cols = []
        return local_cols, ref_table, ref_cols
    return [], '', []

def compare_entities(sql_list, pg_list, entity_type):
    results = []
    matched_pg = set()
    # Entity-specific matching logic
    if entity_type == 'column':
        # Robust column matching: normalize table and column names, and types
        def norm_col(row):
            # Use table name and column name, normalize underscores and case
            tbl = normalize_name(row.get('table',''))
            col = normalize_name(row.get('name',''))
            # Remove underscores for robust matching
            col_nounder = col.replace('_','')
            return (tbl, col_nounder)
        sql_keys = {norm_col(sql): i for i, sql in enumerate(sql_list)}
        pg_keys = {norm_col(pg): i for i, pg in enumerate(pg_list)}
        all_keys = set(sql_keys.keys()) | set(pg_keys.keys())
        matches = {}
        used_pg = set()
        for key in all_keys:
            sql_idx = sql_keys.get(key)
            pg_idx = pg_keys.get(key)
            if sql_idx is not None and pg_idx is not None:
                matches[sql_idx] = pg_idx
                used_pg.add(pg_idx)
        results = []
        matched_pg = set()
        for i, sql in enumerate(sql_list):
            row = {}
            sql_keys_list = sorted([k for k in sql.keys() if not k.startswith('PG_') and k != 'Status'])
            for k in sql_keys_list:
                row['SQL_' + k] = sql[k]
            pg_idx = matches.get(i)
            if pg_idx is not None:
                pg = pg_list[pg_idx]
                matched_pg.add(pg_idx)
                for k in sorted([k for k in pg.keys() if not k.startswith('SQL_') and k != 'Status']):
                    row['PG_' + k] = pg[k]
                # Status logic: compare normalized names and types
                sql_col = normalize_name(sql.get('name','')).replace('_','')
                pg_col = normalize_name(pg.get('name','')).replace('_','')
                if sql_col == pg_col:
                    row['Status'] = 'MATCHED'
                else:
                    row['Status'] = 'MISMATCH: Name variant'
            else:
                row['Status'] = 'MISSING in PG'
            results.append(row)
        # Add unmatched PG
        for i, pg in enumerate(pg_list):
            if i not in matched_pg:
                row = {}
                for k in sorted([k for k in pg.keys() if not k.startswith('SQL_') and k != 'Status']):
                    row['PG_' + k] = pg[k]
                row['Status'] = 'EXTRA in PG'
                results.append(row)
        return results
    elif entity_type == 'table':
        # ...existing code...
        matches = match_by_keys(sql_list, pg_list, ['name'])
    elif entity_type == 'index':
        # ...existing code...
        # ...existing code...
        pass
    elif entity_type == 'constraint':
        # ...existing code...
        pass
    else:
        matches = match_by_keys(sql_list, pg_list, ['name'])
    # SQL-first, row-by-row
    for idx, sql in enumerate(sql_list):
        row = {}
        sql_keys = sorted([k for k in sql.keys() if not k.startswith('PG_') and k != 'Status'])
        for k in sql_keys:
            row[f'SQL_{k}'] = sql.get(k, '')
        pg_idx = matches.get(idx) if entity_type == 'column' else matches.get(id(sql))
        if pg_idx is not None:
            pg = pg_list[pg_idx]
            pg_keys = sorted([k for k in pg.keys() if not k.startswith('SQL_') and k != 'Status'])
            for k in pg_keys:
                row[f'PG_{k}'] = pg.get(k, '')
            matched_pg.add(pg_idx)
            # Status logic
            row['Status'] = 'MATCHED'
        else:
            row['Status'] = 'MISSING in PG'
        results.append(row)
    # Add unmatched PG
    for i, pg in enumerate(pg_list):
        if i not in matched_pg:
            row = {}
            pg_keys = sorted([k for k in pg.keys() if not k.startswith('SQL_') and k != 'Status'])
            for k in pg_keys:
                row[f'PG_{k}'] = pg.get(k, '')
            row['Status'] = 'EXTRA in PG'
            results.append(row)
    return results

def highlight_mismatches(ws):
    status_col = None
    for idx, cell in enumerate(ws[1], 1):
        if str(cell.value).strip().lower() == 'status':
            status_col = idx
            break
    if not status_col:
        return
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    grey_fill = PatternFill(start_color='A9A9A9', end_color='A9A9A9', fill_type='solid')  # Dim grey
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        status = str(row[status_col-1].value or '').lower()
        if 'extra in pg' in status:
            for cell in row:
                cell.fill = grey_fill
        elif any(x in status for x in ['mismatch', 'missing']):
            for cell in row:
                cell.fill = yellow_fill

def write_entity_sheet(wb, sheet_name, compare_rows, columns):
    ws = wb.create_sheet(sheet_name)
    # Reorder columns: SQL_* first, then PG_*, then Difference (if present), then Status
    sql_cols = [c for c in columns if c.startswith('SQL_')]
    pg_cols = [c for c in columns if c.startswith('PG_')]
    diff_cols = [c for c in columns if c == 'Difference']
    other_cols = [c for c in columns if c not in sql_cols + pg_cols + diff_cols + ['Status']]
    out_columns = sql_cols + pg_cols + diff_cols + other_cols + ['Status']
    ws.append(out_columns)
    for row in compare_rows:
        ws.append([row.get(col, '') for col in out_columns])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    if ws.max_row > 1:
        base_table_name = f"Tbl_{sheet_name.replace(' ', '_')}"
        table_name = base_table_name
        existing_table_names = set()
        for sheet in wb.worksheets:
            existing_table_names.update(sheet.tables.keys())
        i = 1
        while table_name in existing_table_names:
            table_name = f"{base_table_name}_{i}"
            i += 1
        table = Table(displayName=table_name, ref=f"A1:{get_column_letter(len(out_columns))}{ws.max_row}")
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length+2, 50)
    highlight_mismatches(ws)

def write_overview_sheet(wb, summary_counts, entity_details=None, db_name=None, server=None, report_date=None):
    from openpyxl.styles import PatternFill, Alignment, Font
    ws = wb.create_sheet('Overview', 0)
    # Title row
    title = f"{db_name or ''} - SCHEMA VALIDATION REPORT"
    ws.merge_cells('A1:F1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    # Server/date row
    ws.merge_cells('A2:C2')
    ws.merge_cells('D2:F2')
    ws['A2'] = f"Server : {server or ''}"
    ws['A2'].font = Font(bold=True)
    ws['A2'].alignment = Alignment(horizontal='left')
    ws['D2'] = f"DATE: {report_date or ''}"
    ws['D2'].font = Font(bold=True)
    ws['D2'].alignment = Alignment(horizontal='right')
    # Header row
    ws.append(['Entity', 'SQL Count', 'PG Count', 'Difference', 'Status', 'Reason'])
    datacounts_totals = None
    # Precompute total SQL/PG counts for DataCounts entity
    if 'DataCounts' in wb.sheetnames:
        dc_ws = wb['DataCounts']
        sql_total = 0
        pg_total = 0
        # Find column indices
        sql_idx = pg_idx = None
        for idx, cell in enumerate(dc_ws[1], 1):
            if str(cell.value).strip().lower() == 'sql_count':
                sql_idx = idx
            if str(cell.value).strip().lower() == 'pg_count':
                pg_idx = idx
        if sql_idx and pg_idx:
            for row in dc_ws.iter_rows(min_row=2, max_row=dc_ws.max_row):
                try:
                    sql_total += int(row[sql_idx-1].value or 0)
                except Exception:
                    pass
                try:
                    pg_total += int(row[pg_idx-1].value or 0)
                except Exception:
                    pass
            datacounts_totals = (sql_total, pg_total)
    for entity, counts in summary_counts.items():
        # For DataCounts, use the sum of SQL_count and PG_count from the DataCounts tab
        if entity == 'DataCounts' and datacounts_totals:
            sql_count = datacounts_totals[0]
            pg_count = datacounts_totals[1]
        else:
            sql_count = counts.get('sql', 0)
            pg_count = counts.get('pg', 0)
        diff = sql_count - pg_count
        status = 'Passed'
        reason_parts = []
        mismatch_rows = []
        missing_rows = []
        extra_rows = []
        # Scan the actual sheet for this entity, if it exists
        sheet_name = entity
        if sheet_name in wb.sheetnames:
            entity_ws = wb[sheet_name]
            # Find the Status column index
            status_col_idx = None
            for idx, cell in enumerate(entity_ws[1], 1):
                if str(cell.value).strip().lower() == 'status':
                    status_col_idx = idx
                    break
            if status_col_idx:
                for row in entity_ws.iter_rows(min_row=2, max_row=entity_ws.max_row):
                    status_val = str(row[status_col_idx-1].value or '').upper()
                    if 'MISMATCH' in status_val:
                        mismatch_rows.append(row)
                    if 'MISSING IN PG' in status_val:
                        missing_rows.append(row)
                    if 'EXTRA IN PG' in status_val:
                        extra_rows.append(row)
        # Status logic: only fail if MISMATCH or MISSING IN PG present (any variant)
        if mismatch_rows or missing_rows:
            status = 'Failed'
        else:
            status = 'Passed'
        # Reason logic
        show_diff_missing = diff > 0 and (not missing_rows or diff != len(missing_rows))
        if show_diff_missing:
            reason_parts.append(f"{diff} missing in PG")
        elif diff > 0 and missing_rows and diff == len(missing_rows):
            # Only show one
            reason_parts.append(f"{len(missing_rows)} missing in PG")
        elif diff < 0:
            reason_parts.append(f"{abs(diff)} extra in PG")
        if missing_rows and not (diff > 0 and diff == len(missing_rows)):
            reason_parts.append(f"{len(missing_rows)} missing in PG")
        if mismatch_rows:
            reason_parts.append(f"{len(mismatch_rows)} mismatches")
        if extra_rows and not (mismatch_rows or missing_rows):
            reason_parts.append(f"{len(extra_rows)} extra in PG")
        if not reason_parts:
            reason = 'All matched'
        else:
            reason = '; '.join(reason_parts)
        ws.append([entity, sql_count, pg_count, diff, status, reason])
    for cell in ws[3]:
        cell.font = Font(bold=True)
    # Color Status column: green for Passed, red for Failed
    status_col_idx = None
    for idx, cell in enumerate(ws[3], 1):
        if str(cell.value).strip().lower() == 'status':
            status_col_idx = idx
            break
    if status_col_idx:
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            status_val = str(row[status_col_idx-1].value or '').strip().lower()
            if status_val == 'passed':
                row[status_col_idx-1].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green
            elif status_val == 'failed':
                row[status_col_idx-1].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Red
    if ws.max_row > 1:
        base_table_name = 'Tbl_Overview'
        table_name = base_table_name
        existing_table_names = set()
        for sheet in wb.worksheets:
            existing_table_names.update(sheet.tables.keys())
        i = 1
        while table_name in existing_table_names:
            table_name = f"{base_table_name}_{i}"
            i += 1
        table = Table(displayName=table_name, ref=f"A3:F{ws.max_row}")
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length+2, 50)

# --- Main ---
def main():
    # Use DB_LIST from config.py for database list
    db_list = DB_LIST
    script_dir = os.path.dirname(os.path.abspath(__file__))
    reports_dir = os.path.join(script_dir, 'SchemaValidationReports')
    os.makedirs(reports_dir, exist_ok=True)
    for db in db_list:
        print(f"\n=== Processing database: {db} ===")
        SQL_SERVER_CONFIG['database'] = db
        POSTGRES_CONFIG['database'] = db
        print("Connecting to SQL Server and PostgreSQL...")
        sql_conn = get_sqlserver_connection()
        pg_conn = get_postgres_connection()
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        summary_counts = {}
        entity_order = [
            ('Tables', 'table', extract_tables),
            ('Columns', 'column', extract_columns),
            # Constraints and Checks will be handled separately
            ('Indexes', 'index', extract_indexes),
            ('Triggers', 'trigger', extract_triggers),
            ('EventTriggers', 'eventtrigger', extract_event_triggers),
            ('Views', 'view', extract_views),
            ('Functions', 'function', extract_functions),  # Ensure Functions is present
            ('Types', 'type', extract_types),
            ('Procedures', 'procedure', extract_procedures),
            ('DataCounts', 'datacounts', extract_table_counts),
        ]
        # --- Extract constraints and split into PK/FK/DEFAULT and CHECK ---
        print(f"\n[Step] Extracting Constraints and Checks...")
        sql_constraints_all = filter_excluded(extract_constraints(sql_conn, 'sql'))
        pg_constraints_all = filter_excluded(extract_constraints(pg_conn, 'pg'))
        # Split
        def is_check(c):
            return normalize_name(c.get('type','')) == 'check'
        def is_pk_fk_default(c):
            t = normalize_name(c.get('type',''))
            return t in ('primary key', 'foreign key', 'default')
        sql_checks = [c for c in sql_constraints_all if is_check(c)]
        pg_checks = [c for c in pg_constraints_all if is_check(c)]
        sql_constraints = [c for c in sql_constraints_all if is_pk_fk_default(c)]
        pg_constraints = [c for c in pg_constraints_all if is_pk_fk_default(c)]
        # --- Constraints Tab (PK, FK, CHECK, DEFAULT) ---
        # [Old logic commented out for reference]
        # print(f"Comparing Constraints (PK, FK, DEFAULT)...")
        # compare_rows = compare_entities(sql_constraints, pg_constraints, 'constraint')
        # summary_counts['Constraints'] = {'sql': len(sql_constraints), 'pg': len(pg_constraints)}
        # all_fields = set()
        # for row in compare_rows:
        #     all_fields.update(row.keys())
        # out_columns = [c for c in sorted(all_fields) if c != 'Status'] + ['Status']
        # print(f"Writing Constraints tab to Excel...")
        # write_entity_sheet(wb, 'Constraints', compare_rows, out_columns)
        # --- CHECKS Tab ---
        # print(f"Comparing CHECKS...")
        # compare_rows = compare_entities(sql_checks, pg_checks, 'constraint')
        # summary_counts['CHECKS'] = {'sql': len(sql_checks), 'pg': len(pg_checks)}
        # all_fields = set()
        # for row in compare_rows:
        #     all_fields.update(row.keys())
        # out_columns = [c for c in sorted(all_fields) if c != 'Status'] + ['Status']
        # print(f"Writing CHECKS tab to Excel...")
        # write_entity_sheet(wb, 'CHECKS', compare_rows, out_columns)

        # --- New Table-wise Constraints Tab (All constraints, schema/table/constraint names/counts) ---
        print(f"Building new table-wise Constraints tab... [NEW LOGIC v2025-09-16]")
        # Get set of base tables (schema, table) for filtering
        sql_base_tables = set((normalize_name(t['schema']), normalize_name(t['name'])) for t in extract_tables(sql_conn, 'sql'))
        pg_base_tables = set((normalize_name(t['schema']), normalize_name(t['name'])) for t in extract_tables(pg_conn, 'pg'))
        def group_constraints_flat(constraints, allowed_tables, dbtype=None):
            grouped = {}
            # Use sets for unique constraint names per type
            type_sets = {}
            for c in constraints:
                schema = normalize_name(c.get('schema',''))
                table = normalize_name(c.get('table',''))
                key = (schema, table)
                if key not in allowed_tables:
                    continue  # Only include base tables
                name = c.get('name','')
                ctype = normalize_name(c.get('type',''))
                # For PG default constraints, append _default for clarity
                if dbtype == 'pg' and ctype == 'default':
                    name = f"{name}_default"
                if key not in grouped:
                    grouped[key] = []
                grouped[key].append(name)
                # Use sets for unique constraint names per type
                if key not in type_sets:
                    type_sets[key] = {'fk': set(), 'pk': set(), 'check': set(), 'default': set()}
                if ctype == 'foreign key':
                    type_sets[key]['fk'].add(name)
                elif ctype == 'primary key':
                    type_sets[key]['pk'].add(name)
                elif ctype == 'check':
                    type_sets[key]['check'].add(name)
                elif ctype == 'default':
                    type_sets[key]['default'].add(name)
            # Convert sets to counts for output
            type_counts = {k: {t: len(v[t]) for t in v} for k, v in type_sets.items()}
            return grouped, type_counts
        sql_grouped, sql_type_counts = group_constraints_flat(sql_constraints_all, sql_base_tables, dbtype='sql')
        pg_grouped, pg_type_counts = group_constraints_flat(pg_constraints_all, pg_base_tables, dbtype='pg')
        all_keys = set(sql_grouped.keys()) | set(pg_grouped.keys())
        compare_rows = []
        for key in sorted(all_keys):
            sql_schema, sql_table = key
            sql_constraints = sql_grouped.get(key, [])
            pg_constraints = pg_grouped.get(key, [])
            # Deduplicate and sort constraint names
            sql_constraints_unique = sorted(set(sql_constraints))
            pg_constraints_unique = sorted(set(pg_constraints))
            # Get type counts for Reason
            sql_types = sql_type_counts.get(key, {'fk':0, 'pk':0, 'check':0, 'default':0})
            pg_types = pg_type_counts.get(key, {'fk':0, 'pk':0, 'check':0, 'default':0})
            reason_parts = []
            for label, typename in [('FK', 'fk'), ('PK', 'pk'), ('Check', 'check'), ('Default', 'default')]:
                diff = sql_types[typename] - pg_types[typename]
                if diff > 0:
                    plural = '' if diff == 1 else 's'
                    reason_parts.append(f"{diff} {label}{plural} is missing")
            reason = ' | '.join(reason_parts)
            row = {
                'sql_schema': sql_schema,
                'sql_tablename': sql_table,
                'sql_constraints': ','.join(sql_constraints_unique),
                'sql_constraints_count': len(sql_constraints_unique),
                'pg_schema': sql_schema,  # Use same key for both, fallback to sql_schema if missing in PG
                'pg_tablename': sql_table,
                'pg_constraints': ','.join(pg_constraints_unique),
                'pg_constraints_count': len(pg_constraints_unique),
                'constraints_logic_version': 'v2025-09-16',  # Marker for new logic
                'Reason': reason
            }
            row['Status'] = 'MATCHED' if row['sql_constraints_count'] == row['pg_constraints_count'] else 'MISMATCH'
            compare_rows.append(row)
        out_columns = [
            'sql_schema', 'sql_tablename', 'sql_constraints', 'sql_constraints_count',
            'pg_schema', 'pg_tablename', 'pg_constraints', 'pg_constraints_count',
            'constraints_logic_version',
            'Reason',
            'Status'
        ]
        print(f"Writing new Constraints tab to Excel... [NEW LOGIC v2025-09-16]")
        write_entity_sheet(wb, 'Constraints', compare_rows, out_columns)
        # Calculate total unique constraints for overview
        total_sql_constraints = sum(len(set(sql_grouped.get(key, []))) for key in sql_grouped)
        total_pg_constraints = sum(len(set(pg_grouped.get(key, []))) for key in pg_grouped)
        summary_counts['Constraints'] = {'sql': total_sql_constraints, 'pg': total_pg_constraints}
        # --- New Table-wise Indexes Tab (All indexes, schema/table/index names/counts) ---
        print(f"Building new table-wise Indexes tab... [NEW LOGIC v2025-09-16]")
        # Get set of base tables (schema, table) for filtering
        sql_base_tables = set((normalize_name(t['schema']), normalize_name(t['name'])) for t in extract_tables(sql_conn, 'sql'))
        pg_base_tables = set((normalize_name(t['schema']), normalize_name(t['name'])) for t in extract_tables(pg_conn, 'pg'))
        def group_indexes_flat(indexes, allowed_tables):
            grouped = {}
            index_defs = {}
            for idx in indexes:
                schema = normalize_name(idx.get('schema',''))
                table = normalize_name(idx.get('table',''))
                key = (schema, table)
                if key not in allowed_tables:
                    continue  # Only include base tables
                name = idx.get('name','')
                columns = idx.get('columns','')
                if key not in grouped:
                    grouped[key] = []
                    index_defs[key] = {}
                grouped[key].append(name)
                index_defs[key][name] = columns
            return grouped, index_defs
        sql_indexes_all = filter_excluded(extract_indexes(sql_conn, 'sql'))
        pg_indexes_all = filter_excluded(extract_indexes(pg_conn, 'pg'))
        sql_grouped_idx, sql_index_defs = group_indexes_flat(sql_indexes_all, sql_base_tables)
        pg_grouped_idx, pg_index_defs = group_indexes_flat(pg_indexes_all, pg_base_tables)
        all_idx_keys = set(sql_grouped_idx.keys()) | set(pg_grouped_idx.keys())
        index_compare_rows = []
        for key in sorted(all_idx_keys):
            sql_schema, sql_table = key
            sql_indexes = sql_grouped_idx.get(key, [])
            pg_indexes = pg_grouped_idx.get(key, [])
            # Deduplicate and sort index names
            sql_indexes_unique = sorted(set(sql_indexes))
            pg_indexes_unique = sorted(set(pg_indexes))
            row = {
                'sql_schema': sql_schema,
                'sql_tablename': sql_table,
                'sql_indexes': ','.join(sql_indexes_unique),
                'sql_indexes_count': len(sql_indexes_unique),
                'pg_schema': sql_schema,
                'pg_tablename': sql_table,
                'pg_indexes': ','.join(pg_indexes_unique),
                'pg_indexes_count': len(pg_indexes_unique),
            }
            # Status logic: if SQL count is 0 and PG count > 0, mark as EXTRA in PG
            if row['sql_indexes_count'] == 0 and row['pg_indexes_count'] > 0:
                row['Status'] = 'EXTRA in PG'
            else:
                row['Status'] = 'MATCHED' if row['sql_indexes_count'] == row['pg_indexes_count'] else 'MISMATCH'
            index_compare_rows.append(row)
        out_columns = [
            'sql_schema', 'sql_tablename', 'sql_indexes', 'sql_indexes_count',
            'pg_schema', 'pg_tablename', 'pg_indexes', 'pg_indexes_count',
            'Reason',
            'Status'
        ]
        print(f"Writing new Indexes tab to Excel... [NEW LOGIC v2025-09-16]")
        write_entity_sheet(wb, 'Indexes', index_compare_rows, out_columns)
        # Calculate total unique indexes for overview
        total_sql_indexes = sum(len(set(sql_grouped_idx.get(key, []))) for key in sql_grouped_idx)
        total_pg_indexes = sum(len(set(pg_grouped_idx.get(key, []))) for key in pg_grouped_idx)
        summary_counts['Indexes'] = {'sql': total_sql_indexes, 'pg': total_pg_indexes}
        # --- New Table-wise Triggers Tab (All triggers, schema/table/trigger names/counts) ---
        print(f"Building new table-wise Triggers tab... [NEW LOGIC v2025-09-16]")
        # Get set of base tables (schema, table) for filtering
        sql_base_tables = set((normalize_name(t['schema']), normalize_name(t['name'])) for t in extract_tables(sql_conn, 'sql'))
        pg_base_tables = set((normalize_name(t['schema']), normalize_name(t['name'])) for t in extract_tables(pg_conn, 'pg'))
        def group_triggers_flat(triggers, allowed_tables):
            grouped = {}
            for tr in triggers:
                schema = normalize_name(tr.get('schema',''))
                table = normalize_name(tr.get('table',''))
                key = (schema, table)
                # Fix: If allowed_tables is empty, allow all; else, check membership
                if allowed_tables and key not in allowed_tables:
                    continue  # Only include base tables if specified
                name = tr.get('name','')
                if key not in grouped:
                    grouped[key] = []
                # Fix: Always append, do not deduplicate here (deduplication is done later)
                grouped[key].append(name)
            return grouped
        sql_triggers_all = filter_excluded(extract_triggers(sql_conn, 'sql'))
        pg_triggers_all = filter_excluded(extract_triggers(pg_conn, 'pg'))
        sql_grouped_tr = group_triggers_flat(sql_triggers_all, sql_base_tables)
        pg_grouped_tr = group_triggers_flat(pg_triggers_all, pg_base_tables)
        all_tr_keys = set(sql_grouped_tr.keys()) | set(pg_grouped_tr.keys())
        trigger_compare_rows = []
        def strip_pg_event_suffix(name):
            # Remove _insert, _update, _delete suffixes for robust matching
            return re.sub(r'_(insert|update|delete)$', '', name, flags=re.IGNORECASE)
        def norm_trigger_name(name):
            # Lowercase, remove underscores for robust matching
            return (name or '').replace('_','').lower()
        for key in sorted(all_tr_keys):
            sql_schema, sql_table = key
            sql_triggers = sql_grouped_tr.get(key, [])
            pg_triggers = pg_grouped_tr.get(key, [])
            sql_triggers_unique = sorted(set(sql_triggers))
            pg_triggers_unique = sorted(set(pg_triggers))
            sql_norm = [norm_trigger_name(t) for t in sql_triggers_unique]
            pg_norm = [norm_trigger_name(t) for t in pg_triggers_unique]
            sql_bases = [strip_pg_event_suffix(s) for s in sql_norm]
            pg_bases = [strip_pg_event_suffix(p) for p in pg_norm]
            # Robust matching: consider prefix match for truncation
            missing_pg = []
            for i, s in enumerate(sql_bases):
                found = False
                for p in pg_bases:
                    if s == p or s.startswith(p) or p.startswith(s):
                        found = True
                        break
                if not found:
                    missing_pg.append(sql_triggers_unique[i])
            extra_pg = []
            for j, p in enumerate(pg_bases):
                found = False
                for s in sql_bases:
                    if s == p or s.startswith(p) or p.startswith(s):
                        found = True
                        break
                if not found:
                    extra_pg.append(pg_triggers_unique[j])
            row = {
                'sql_schema': sql_schema,
                'sql_tablename': sql_table,
                'sql_triggers': ','.join(sql_triggers_unique),
                'sql_triggers_count': len(sql_triggers_unique),
                'pg_schema': sql_schema,
                'pg_tablename': sql_table,
                'pg_triggers': ','.join(pg_triggers_unique),
                'pg_triggers_count': len(pg_triggers_unique),
            }
            reason_parts = []
            if len(missing_pg) > 0:
                plural = '' if len(missing_pg) == 1 else 's'
                reason_parts.append(f"Missing in PG: {','.join(missing_pg)}")
            if len(extra_pg) > 0:
                plural = '' if len(extra_pg) == 1 else 's'
                reason_parts.append(f"Extra in PG: {','.join(extra_pg)}")
            # Status logic: MATCHED if all SQL bases in PG and all PG bases in SQL
            if row['sql_triggers_count'] == 0 and row['pg_triggers_count'] > 0:
                row['Status'] = 'EXTRA in PG'
            elif not missing_pg and not extra_pg:
                row['Status'] = 'MATCHED'
            else:
                row['Status'] = 'MISMATCH'
            trigger_compare_rows.append(row)
        out_columns = [
            'sql_schema', 'sql_tablename', 'sql_triggers', 'sql_triggers_count',
            'pg_schema', 'pg_tablename', 'pg_triggers', 'pg_triggers_count',
            'Reason',
            'Status'
        ]
        print(f"Writing new Triggers tab to Excel... [NEW LOGIC v2025-09-16]")
        write_entity_sheet(wb, 'Triggers', trigger_compare_rows, out_columns)
        # Calculate total unique triggers for overview
        total_sql_triggers = sum(len(set(sql_grouped_tr.get(key, []))) for key in sql_grouped_tr)
        total_pg_triggers = sum(len(set(pg_grouped_tr.get(key, []))) for key in pg_grouped_tr)
        summary_counts['Triggers'] = {'sql': total_sql_triggers, 'pg': total_pg_triggers}
        # --- Improved EventTriggers Tab with name mapping ---
        print(f"\n[Step] Extracting EventTriggers with mapping...")
        sql_event_triggers = filter_excluded(extract_event_triggers(sql_conn, 'sql'))
        pg_event_triggers = filter_excluded(extract_event_triggers(pg_conn, 'pg'))
        sql_names = [et['name'] for et in sql_event_triggers]
        pg_names = [et['name'] for et in pg_event_triggers]
        pg_types = [et.get('event_type', et.get('type', '')) for et in pg_event_triggers]  # dynamic event type
        matched_pg = set()
        compare_rows = []
        for sql_et in sql_event_triggers:
            sql_name = sql_et['name']
            sql_type = sql_et.get('event_type', sql_et.get('type', '')) or 'trigger'
            mapped_pg_names = EVENT_TRIGGER_NAME_MAP.get(sql_name, [])
            found_pg = []
            found_pg_types = []
            for mapped_pg in mapped_pg_names:
                for i, pg_name in enumerate(pg_names):
                    if normalize_name(pg_name) == normalize_name(mapped_pg):
                        found_pg.append(pg_name)
                        found_pg_types.append(pg_types[i] or 'event_trigger')
                        matched_pg.add(i)
            if mapped_pg_names and found_pg:
                # Mapped and found
                row = {"SQL_name": sql_name, "SQL_event_type": sql_type, "PG_name": ','.join(found_pg), "PG_event_type": ','.join(found_pg_types), "Status": "MATCHED", "Reason": "Mapped and found in PG"}
            elif mapped_pg_names:
                # Mapped but not found
                row = {"SQL_name": sql_name, "SQL_event_type": sql_type, "PG_name": ','.join(mapped_pg_names), "PG_event_type": '', "Status": "MISSING in PG", "Reason": "Mapped PG event trigger(s) not found"}
            else:
                # Fallback to normalized name matching
                found = False
                for i, pg_name in enumerate(pg_names):
                    if normalize_name(sql_name) == normalize_name(pg_name):
                        row = {"SQL_name": sql_name, "SQL_event_type": sql_type, "PG_name": pg_name, "PG_event_type": pg_types[i] or 'event_trigger', "Status": "MATCHED", "Reason": "Direct name match"}
                        matched_pg.add(i)
                        found = True
                        break
                if not found:
                    row = {"SQL_name": sql_name, "SQL_event_type": sql_type, "PG_name": '', "PG_event_type": '', "Status": "MISSING in PG", "Reason": "No matching event trigger in PG"}
            compare_rows.append(row)
        # Add unmatched PG event triggers
        for i, pg_name in enumerate(pg_names):
            if i not in matched_pg:
                # Check if this PG name is in any mapped list
                mapped = False
                for mapped_list in EVENT_TRIGGER_NAME_MAP.values():
                    if normalize_name(pg_name) in [normalize_name(x) for x in mapped_list]:
                        mapped = True
                        break
                if not mapped:
                    row = {"SQL_name": '', "SQL_event_type": '', "PG_name": pg_name, "PG_event_type": pg_types[i] or 'event_trigger', "Status": "EXTRA in PG", "Reason": "Extra event trigger in PG"}
                    compare_rows.append(row)
        out_columns = ["SQL_name", "SQL_event_type", "PG_name", "PG_event_type", "Status", "Reason"]
        print(f"Writing EventTriggers tab to Excel... [IMPROVED LOGIC]")
        write_entity_sheet(wb, "EventTriggers", compare_rows, out_columns)
        summary_counts["EventTriggers"] = {"sql": len(sql_event_triggers), "pg": len(pg_event_triggers)}
        # --- End Improved EventTriggers Tab ---
        # --- Improved Procedures Tab with mapping and robust row alignment (EventTriggers logic) ---
        print(f"\n[Step] Extracting Procedures with mapping...")
        sql_procs = filter_excluded(extract_procedures(sql_conn, 'sql'))
        pg_procs = filter_excluded(extract_procedures(pg_conn, 'pg'))
        sql_proc_names = [p['name'] for p in sql_procs]
        pg_proc_names = [p['name'] for p in pg_procs]
        matched_pg = set()
        compare_rows = []
        for sql_proc in sql_procs:
            sql_name = sql_proc['name']
            mapped_pg_names = PROCEDURE_NAME_MAP.get(sql_name, [])
            found_pg = []
            for mapped_pg in mapped_pg_names:
                for i, pg_name in enumerate(pg_proc_names):
                    if normalize_name(pg_name) == normalize_name(mapped_pg):
                        found_pg.append(pg_name)
                        matched_pg.add(i)
            if mapped_pg_names and found_pg:
                row = {"SQL_name": sql_name, "PG_name": ','.join(found_pg), "Status": "MATCHED", "Reason": "Mapped and found in PG"}
            elif mapped_pg_names:
                row = {"SQL_name": sql_name, "PG_name": ','.join(mapped_pg_names), "Status": "MISSING in PG", "Reason": "Mapped PG procedure(s) not found"}
            else:
                found = False
                for i, pg_name in enumerate(pg_proc_names):
                    if normalize_name(sql_name) == normalize_name(pg_name):
                        row = {"SQL_name": sql_name, "PG_name": pg_name, "Status": "MATCHED", "Reason": "Direct name match"}
                        matched_pg.add(i)
                        found = True
                        break
                if not found:
                    row = {"SQL_name": sql_name, "PG_name": '', "Status": "MISSING in PG", "Reason": "No matching procedure in PG"}
            compare_rows.append(row)
        for i, pg_name in enumerate(pg_proc_names):
            if i not in matched_pg:
                mapped = False
                for mapped_list in PROCEDURE_NAME_MAP.values():
                    if normalize_name(pg_name) in [normalize_name(x) for x in mapped_list]:
                        mapped = True
                        break
                if not mapped:
                    row = {"SQL_name": '', "PG_name": pg_name, "Status": "EXTRA in PG", "Reason": "Extra procedure in PG"}
                    compare_rows.append(row)
        out_columns = ["SQL_name", "PG_name", "Status", "Reason"]
        print(f"Writing Procedures tab to Excel... [IMPROVED LOGIC]")
        write_entity_sheet(wb, "Procedures", compare_rows, out_columns)
        summary_counts["Procedures"] = {"sql": len(sql_procs), "pg": len(pg_procs)}
        # Remove Procedures1 sheet if it exists (Excel may auto-create it if duplicate names)
        if 'Procedures1' in wb.sheetnames:
            std = wb['Procedures1']
            wb.remove(std)
        # --- Rest of the tabs ---
        entity_details = {}  # Collect details for overview
        for sheet, entity_type, extractor in entity_order:
            print(f"\n[Step] Extracting {sheet}...")
            if sheet in ('Constraints', 'CHECKS', 'Indexes', 'Triggers', 'EventTriggers', 'Procedures'):
                continue  # Already handled or handled specially
            if sheet == 'Types':
                print("Comparing Types with robust/fuzzy matching and SQL/PG columns...")
                sql_types = extract_types(sql_conn, 'sql')
                pg_types = extract_types(pg_conn, 'pg')
                def norm_type_name(name):
                    return (name or '').replace('_', '').lower()
                matched_pg = set()
                compare_rows = []
                for sql in sql_types:
                    sql_name = norm_type_name(sql['type_name'])
                    sql_kind = sql.get('type_kind', '')
                    best_pg = None
                    for i, pg in enumerate(pg_types):
                        if i in matched_pg:
                            continue
                        pg_name = norm_type_name(pg['type_name'])
                        if sql_name == pg_name or sql_name in pg_name or pg_name in sql_name:
                            best_pg = i
                            break
                    row = {
                        'SQL_schema': sql['schema'],
                        'SQL_type_name': sql['type_name'],
                        'SQL_type_kind': sql_kind,
                    }
                    if best_pg is not None:
                        pg = pg_types[best_pg]
                        row['PG_schema'] = pg['schema']
                        row['PG_type_name'] = pg['type_name']
                        row['PG_type_kind'] = pg.get('type_kind', '')
                        row['Reason'] = ''
                        row['Status'] = 'MATCHED'
                        matched_pg.add(best_pg)
                    else:
                        row['PG_schema'] = ''
                        row['PG_type_name'] = ''
                        row['PG_type_kind'] = ''
                        row['Reason'] = 'Missing in PG'
                        row['Status'] = 'MISSING in PG'
                    compare_rows.append(row)
                # Add unmatched PG types
                for i, pg in enumerate(pg_types):
                    if i not in matched_pg:
                        row = {
                            'SQL_schema': '',
                            'SQL_type_name': '',
                            'SQL_type_kind': '',
                            'PG_schema': pg['schema'],
                            'PG_type_name': pg['type_name'],
                            'PG_type_kind': pg.get('type_kind', ''),
                            'Reason': 'Extra in PG',
                            'Status': 'EXTRA in PG'
                        }
                        compare_rows.append(row)
                out_columns = ['SQL_schema', 'SQL_type_name', 'SQL_type_kind', 'PG_schema', 'PG_type_name', 'PG_type_kind', 'Reason', 'Status']
                print("Writing Types tab to Excel... [ROBUST LOGIC]")
                write_entity_sheet(wb, 'Types', compare_rows, out_columns)
                summary_counts['Types'] = {'sql': len(sql_types), 'pg': len(pg_types)}
                continue
            if sheet == 'DataCounts':
                continue  # Improved logic below
            if sheet == 'Functions':
                # --- Functions Tab: Only normal functions (exclude trigger functions) ---
                print(f"Building Functions tab (excluding trigger functions)...")
                sql_functions_all = filter_excluded(extract_functions(sql_conn, 'sql'))
                pg_functions_all = filter_excluded(extract_functions(pg_conn, 'pg'))
                sql_normal_functions = [f for f in sql_functions_all if f.get('function_type', 'normal') == 'normal']
                pg_normal_functions = [f for f in pg_functions_all if f.get('function_type', 'normal') == 'normal']
                compare_rows = compare_entities(sql_normal_functions, pg_normal_functions, 'function')
                # Additional schema name mismatch check and status/Reason logic
                for row in compare_rows:
                    sql_name = row.get('SQL_name') or row.get('SQL_name', '')
                    pg_name = row.get('PG_name') or row.get('PG_name', '')
                    sql_schema = row.get('SQL_schema', '')
                    pg_schema = row.get('PG_schema', '')
                    status = row.get('Status', '')
                    # If matched and schema also matches, set Status to 'MATCHED' and Reason to ''
                    if sql_name and pg_name and normalize_name(sql_name) == normalize_name(pg_name):
                        if sql_schema and pg_schema and normalize_name(sql_schema) != normalize_name(pg_schema):
                            row['Status'] = 'MISMATCH'
                            row['Reason'] = 'Schema name mismatch'
                        elif status.startswith('MATCHED'):
                            row['Status'] = 'MATCHED'
                            row['Reason'] = ''
                    elif status.startswith('MATCHED'):
                        row['Status'] = 'MATCHED'
                        row['Reason'] = ''
                    elif status.startswith('MISMATCH') and not row.get('Reason'):
                        row['Reason'] = 'Name mismatch'
                    elif status.startswith('MISSING') and not row.get('Reason'):
                        row['Reason'] = 'Missing in PG'
                    elif status.startswith('EXTRA') and not row.get('Reason'):
                        row['Reason'] = 'Extra in PG'
                all_fields = set()
                for row in compare_rows:
                    all_fields.update(row.keys())
                out_columns = [c for c in sorted(all_fields) if c not in ('Status','Reason')] + ['Reason','Status']
                write_entity_sheet(wb, 'Functions', compare_rows, out_columns)
                summary_counts['Functions'] = {'sql': len(sql_normal_functions), 'pg': len(pg_normal_functions)}

                # --- Trigger Functions Tab: Only trigger functions from dbo, meta, public schemas ---
                print(f"Building Trigger Functions tab (trigger functions from dbo/meta/public)...")
                allowed_schemas = {'dbo', 'meta', 'public'}
                def is_allowed_schema(f):
                    return normalize_name(f.get('schema','')) in allowed_schemas
                sql_trigger_functions = [f for f in sql_functions_all if f.get('function_type', 'normal') == 'trigger' and is_allowed_schema(f)]
                pg_trigger_functions = [f for f in pg_functions_all if f.get('function_type', 'normal') == 'trigger' and is_allowed_schema(f)]
                compare_rows = compare_entities(sql_trigger_functions, pg_trigger_functions, 'function')
                # Additional schema name mismatch check and status/Reason logic
                for row in compare_rows:
                    sql_name = row.get('SQL_name') or row.get('SQL_name', '')
                    pg_name = row.get('PG_name') or row.get('PG_name', '')
                    sql_schema = row.get('SQL_schema', '')
                    pg_schema = row.get('PG_schema', '')
                    status = row.get('Status', '')
                    if sql_name and pg_name and normalize_name(sql_name) == normalize_name(pg_name):
                        if sql_schema and pg_schema and normalize_name(sql_schema) != normalize_name(pg_schema):
                            row['Status'] = 'MISMATCH'
                            row['Reason'] = 'Schema name mismatch'
                        elif status.startswith('MATCHED'):
                            row['Status'] = 'MATCHED'
                            row['Reason'] = ''
                    elif status.startswith('MATCHED'):
                        row['Status'] = 'MATCHED'
                        row['Reason'] = ''
                    elif status.startswith('MISMATCH') and not row.get('Reason'):
                        row['Reason'] = 'Name mismatch'
                    elif status.startswith('MISSING') and not row.get('Reason'):
                        row['Reason'] = 'Missing in PG'
                    elif status.startswith('EXTRA') and not row.get('Reason'):
                        row['Reason'] = 'Extra in PG'
                all_fields = set()
                for row in compare_rows:
                    all_fields.update(row.keys())
                out_columns = [c for c in sorted(all_fields) if c not in ('Status','Reason')] + ['Reason','Status']
                write_entity_sheet(wb, 'Trigger Functions', compare_rows, out_columns)
                summary_counts['Trigger Functions'] = {'sql': len(sql_trigger_functions), 'pg': len(pg_trigger_functions)}
                continue
            print(f"Extracting {sheet}...")
            sql_data = filter_excluded(extractor(sql_conn, 'sql')) if sheet != 'EventTriggers' else []
            pg_data = filter_excluded(extractor(pg_conn, 'pg'))
            print(f"Comparing {sheet}...")
            compare_rows = compare_entities(sql_data, pg_data, entity_type)
            summary_counts[sheet] = {
                'sql': len(sql_data) if sheet != 'EventTriggers' else 0,
                'pg': len(pg_data)
            }
            all_fields = set()
            for row in compare_rows:
                all_fields.update(row.keys())
            out_columns = [c for c in sorted(all_fields) if c != 'Status'] + ['Status']
            print(f" Writing {sheet} tab to Excel...")
            write_entity_sheet(wb, sheet, compare_rows, out_columns)
        # --- Improved DataCounts Tab ---
        print(f"\n[Step] Extracting DataCounts with schema/table/percentage match...")
        sql_counts = filter_excluded(extract_table_counts(sql_conn, 'sql'))
        pg_counts = filter_excluded(extract_table_counts(pg_conn, 'pg'))
        def norm_schema_table(row):
            return (normalize_name(row.get('schema','')), normalize_name(row.get('name','')))
        sql_lookup = {norm_schema_table(row): row for row in sql_counts}
        pg_lookup = {norm_schema_table(row): row for row in pg_counts}
        all_keys = set(sql_lookup.keys()) | set(pg_lookup.keys())
        compare_rows = []
        for key in sorted(all_keys):
            sql_row = sql_lookup.get(key)
            pg_row = pg_lookup.get(key)
            sql_schema, sql_table = key
            sql_count = int(sql_row['count']) if sql_row and str(sql_row.get('count','')).isdigit() else 0
            pg_count = int(pg_row['count']) if pg_row and str(pg_row.get('count','')).isdigit() else 0
            row = {
                'SQL_schema': sql_schema,
                'SQL_table': sql_table,
                'PG_schema': pg_row['schema'] if pg_row else '',
                'PG_table': pg_row['name'] if pg_row else '',
                'SQL_count': sql_count,
                'PG_count': pg_count,
            }
            if sql_count == pg_count:
                row['Status'] = 'MATCHED'
            elif sql_count == 0 and pg_count == 0:
                row['Status'] = 'MATCHED (both zero)'
            else:
                percent = 0
                if sql_count > 0 and pg_count > 0:
                    percent = int((min(sql_count, pg_count) / max(sql_count, pg_count)) * 100)
                row['Status'] = f"MISMATCH: {percent}% match (SQL: {sql_count}, PG: {pg_count})"
            compare_rows.append(row)
        out_columns = ['SQL_schema', 'SQL_table', 'PG_schema', 'PG_table', 'SQL_count', 'PG_count', 'Status']
        print(f"Writing DataCounts tab to Excel... [IMPROVED LOGIC]")
        if 'DataCounts' in wb.sheetnames:
            std = wb['DataCounts']
            wb.remove(std)
        write_entity_sheet(wb, 'DataCounts', compare_rows, out_columns)
        summary_counts['DataCounts'] = {'sql': len(sql_counts), 'pg': len(pg_counts)}
        # --- Overview Tab (already handled in the original code) ---
        print("Writing Overview tab to Excel...")
        # Pass db, server, and date to write_overview_sheet
        now = datetime.datetime.now().strftime('%d-%m-%Y')
        write_overview_sheet(wb, summary_counts, entity_details, db_name=db, server=SQL_SERVER_CONFIG['server'], report_date=now)
        now_file = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M')
        server = SQL_SERVER_CONFIG['server']
        filename = f'{server}_{db}_Schema_Validation_{now_file}.xlsx'
        file_path = os.path.join(reports_dir, filename)
        print(f"Saving Excel file: {file_path}")
        wb.save(file_path)
        print(f'Validation Excel generated for {db} at {file_path}.')
if __name__ == '__main__':
    main()
